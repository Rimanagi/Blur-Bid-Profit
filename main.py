# pip install -r requirements
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import linecache
import webbrowser

#chrome_path = 'C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s'
chrome_path = 'C:/Program Files/Google/Chrome/Application/chrome.exe %s'

# input here .txt filename
filename = input("Enter .txt filename: ") + ".txt"
input_txt_file = open(filename, 'r')
# input here .xlsx filename
output_xlsx_file = load_workbook(input("Enter .xlsx filename: ") + ".xlsx")
ws = output_xlsx_file.active # for xl sheet


row = 2
number_of_line = 0
blur_bid_word = "Blur Bid Profit: "
value_to_compare = float(input("Минимальное желаемое значение Blur Bid: "))


for line in input_txt_file:
    number_of_line += 1

    if blur_bid_word in line:
        blur_bid_profit = float(line.split()[-1])

        if float(blur_bid_profit) >= value_to_compare:
            #adding Blur Bid in xlsx
            line = linecache.getline(filename, number_of_line - 1)
            blur_bid_value = float(line.split()[-1])
            ws.cell(row=row, column=5).value = blur_bid_value
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="61f551")

            #opening link in Google
            line = linecache.getline(filename, number_of_line - 2)
            if "https" not in line.split()[-1]:
                line = linecache.getline(filename, number_of_line - 3)
            url = str(line.split()[-1])
            webbrowser.get(chrome_path).open(url)
        else:
            break
        row += 1

linecache.clearcache()
input_txt_file.close()
output_xlsx_file.save("2.xlsx")